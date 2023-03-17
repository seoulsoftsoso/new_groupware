import datetime
import math
import os
import shutil

import openpyxl
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.db import transaction
from django.http import JsonResponse
from django.shortcuts import render, redirect
from django.urls import reverse
from django.views.generic import ListView, UpdateView, DeleteView
from openpyxl import load_workbook

import order_manage
from Pagenation import PaginatorManager
from order_manage.form import *
from order_manage.models import *
from seoulsoft_mes.settings import BASE_DIR


class LabelPrint_productAutoComplete(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        qs = LabelPrint_product.objects.filter(
            enterprise__name=self.request.COOKIES['enterprise_name'],
            enable=False
        )

        return qs
class LabelPrint_deliveryAutoComplete(autocomplete.Select2QuerySetView):
    def get_queryset(self):
        qs = LabelPrint_delivery.objects.filter(
            enterprise__name=self.request.COOKIES['enterprise_name'],
            enable=False
        )

        return qs

class LabelPrint(ListView):

    def get(self, request, *args, **kwargs):
        context = {}

        now = datetime.datetime.now()
        context['search_date'] = self.request.GET.get('search_date', now.strftime('%Y-%m-%d'))
        context['search_product'] = self.request.GET.get('search_product', '')
        context['search_customer'] = self.request.GET.get('search_customer', '')
        context['search_customer_name'] = self.request.GET.get('search_customer_name', '')
        # 세팅된 항목 저장
        if context['search_customer'] != '':
            context['search_customer_name'] = LabelPrint_delivery.objects.get(pk=context['search_customer']).delivery_to

        context['searchForm'] = Search_labelprint(request.GET, request.COOKIES['enterprise_name'])
        context['form'] = LabelPrint_History_CreateForm(request.GET, request.COOKIES['enterprise_name'])

        # 쿼리
        qs = LabelPrint_History.objects.none()
        if context['search_customer'] != '':
            qs = LabelPrint_History.objects.filter(
                enterprise__name=self.request.COOKIES['enterprise_name'],
                date=context['search_date'],
                customer_name=context['search_customer_name']
            )
            if context['search_product'] != '':
                qs = qs.filter(product_name__contains=context['search_product'])

        context['count'] = qs.count()

        # 페이지 네이션
        context['page'] = self.request.GET.get('page', 1)
        context['paginate_by'] = self.request.GET.get('paginate_by', 20)
        context["page_range"], context["queryset"] = PaginatorManager(self.request, qs, 20)

        context['method'] = '추가'

        return render(request, 'label_print.html', context)

    def post(self, request):

        form = LabelPrint_History_CreateForm(request.POST, request.COOKIES['enterprise_name'])

        if form.is_valid():
            enterprise = EnterpriseMaster.objects.get(name=request.COOKIES['enterprise_name'])
            user = UserMaster.objects.get(username=request.COOKIES['username'])
            form.instance.enterprise = enterprise
            form.instance.enable = False

            today = datetime.date.today()
            form.instance.created_at = today
            form.instance.updated_at = today
            form.instance.created_by = user
            form.instance.updated_by = user
            form.save()
        else:
            print(form.errors)

        return self.get(request)

class LabelPrint_Update(UpdateView):
    model = LabelPrint_History
    form_class = LabelPrint_History_CreateForm
    template_name = 'label_print.html'

    def get_success_url(self):
        now = datetime.datetime.now()
        search_date = self.request.GET.get('search_date', now.strftime('%Y-%m-%d'))
        search_product = self.request.GET.get('search_product', '')
        search_customer = self.request.GET.get('search_customer', '')
        search_customer_name = self.request.GET.get('search_customer_name', '')
        page = self.request.GET.get('page', 1)

        url = "/ordering/label_print/?page=" + page + "&search_date=" + search_date + "&search_product=" + search_product + '&search_customer=' + search_customer + '&search_customer_name=' + search_customer_name
        return url

    def get_context_data(self, **kwargs):
        context = super(LabelPrint_Update, self).get_context_data(**kwargs)

        now = datetime.datetime.now()
        context['search_date'] = self.request.GET.get('search_date', now.strftime('%Y-%m-%d'))
        context['search_product'] = self.request.GET.get('search_product', '')
        context['search_customer'] = self.request.GET.get('search_customer', '')
        context['search_customer_name'] = self.request.GET.get('search_customer_name', '')
        # 세팅된 항목 저장
        if context['search_customer'] != '':
            context['search_customer_name'] = LabelPrint_delivery.objects.get(pk=context['search_customer']).delivery_to

        context['searchForm'] = Search_labelprint(self.request.GET, self.request.COOKIES['enterprise_name'])

        # 쿼리
        qs = LabelPrint_History.objects.none()
        if context['search_customer'] != '':
            qs = LabelPrint_History.objects.filter(
                enterprise__name=self.request.COOKIES['enterprise_name'],
                date=context['search_date'],
                customer_name=context['search_customer_name']
            )
            if context['search_product'] != '':
                qs = qs.filter(product_name__contains=context['search_product'])

        context['itemId'] = self.kwargs['pk']
        context['count'] = qs.count()

        # 페이지 네이션
        context['page'] = self.request.GET.get('page', 1)
        context['paginate_by'] = self.request.GET.get('paginate_by', 20)
        context["page_range"], context["queryset"] = PaginatorManager(self.request, qs, 20)

        context['method'] = '수정'

        return context

    def form_valid(self, form):
        respons = super(LabelPrint_Update, self).form_valid(form)
        with transaction.atomic():
            self.object = form.save()
            updated = form.save()
            updated.updated_at = datetime.date.today()

            user = UserMaster.objects.get(username=self.request.COOKIES['username'])
            updated.updated_by = user

            updated.save()

        return respons

    def form_invalid(self, form):
        response = super(LabelPrint_Update, self).form_invalid(form)
        print(form.errors)
        return response


class LabelPrint_DeleteView(DeleteView):
    model = LabelPrint_History
    template_name = 'label_print.html'

    def get(self, request, pk):
        try:
            item = LabelPrint_History.objects.get(pk=pk)
            item.delete()
        except:
            pass

        return LabelPrint.get(self, request=request)


def LabelPrintExcel(request):
    context = {}

    try:
        file = request.FILES['file']
    except:
        file = None
    date = request.POST['date']
    customer = request.POST['customer']
    product = request.POST.get('product', None)
    if product == 'null':
        product = None
    remarks = request.POST['remarks']
    enterprise_name = request.POST['enterprise_name']

    user = UserMaster.objects.get(username=request.POST['username'])

    print(date, customer, remarks, enterprise_name)
    print(file)

    # 각 파일이 존재하면 각각의 방식으로 파일 처리
    if file is not None:
        cnt, error = ExcelUpload(file, date, enterprise_name, customer, product, user, remarks)
        context['message'] = str(cnt) + '건 업로드가 완료되었습니다.\n'
        if len(error) > 0:
            context['message'] += error[0]['error']
    else:
        context['message'] = '파일이 없습니다.'

    return JsonResponse(context)


def ExcelUpload(csv_file: InMemoryUploadedFile, date, enterprise, customer, product, user, remarks):
    wb = load_workbook(csv_file, data_only=True)

    # 제품 쿼리 가져오기
    qs_product = LabelPrint_product.objects.filter(enterprise__name=enterprise, enable=False)

    enterprise_obj = EnterpriseMaster.objects.get(name=enterprise)

    cnt = 0
    error_contents = []

    for work_sheet in wb:
        if len(error_contents) > 0:
            break
        for row in work_sheet:
            if len(error_contents) > 0:
                break

            if row[0].value == 'No':
                continue
            else:
                if row[5].value:
                    if row[5].value in enterprise:
                        pass
                    else:
                        error_contents.append({'error': '공급업체 오류 - 타 사의 발주정보가 포함되어 있습니다. ' + row[5].value + '\n'})
                        break
                # 식재료명
                if row[7].value:
                    if '김치' in str(row[7].value):
                        productName = row[7].value[3:len(row[7].value)-1]

                        # 예외처리
                        if '맛김치' in productName:
                            productName = '맛김치'
                        elif '겉절이' in productName:
                            productName = '겉절이'
                        elif '배추김치' in productName:
                            productName = '배추포기김치'
                        elif '총각김치' in productName:
                            productName = '알타리'

                        if qs_product.filter(name__contains=productName).count() > 0:
                            pass
                        else:
                            error_contents.append({'error': '제품 오류 - ' + row[7].value + '에 해당하는 제품 정보 추가가 필요합니다.\n'})
                            break

    print('start')
    startedTime = datetime.datetime.now()
    # 오류가 있어도 일단 생성한다.
    # 오류 항목은 빈 칸으로 처리
    # 기존 데이터 삭제
    product_current = None
    customer_current = LabelPrint_delivery.objects.get(pk=customer)
    qs_history = LabelPrint_History.objects.filter(date=date, customer_name=customer_current.delivery_to, enterprise__name=enterprise)
    if product:
        product_current = LabelPrint_product.objects.get(pk=product)
        qs_history = qs_history.filter(product_name=product_current.name)
    qs_history.delete()

    now = datetime.datetime.now()

    for work_sheet in wb:
        list = []
        append = list.append
        for row in work_sheet:

            if row[0].value == 'No':
                continue
            elif row[4].value is None:
                # 마지막 열 처리
                continue
            else:
                product_current_check = None
                # 제품
                if row[7].value:
                    if '김치' in str(row[7].value):
                        productName = row[7].value[3:len(row[7].value)-1]

                        # 예외처리
                        if '맛김치' in productName:
                            productName = '맛김치'
                        elif '겉절이' in productName:
                            productName = '겉절이'
                        elif '배추김치' in productName:
                            productName = '배추포기김치'
                        elif '총각김치' in productName:
                            productName = '알타리'
                        qs_currentPD = qs_product.filter(name__contains=productName)
                        if qs_currentPD.count() > 0:
                            product_current_check = qs_currentPD.first()

                if product:
                    if product_current != product_current_check:
                        continue

                # 배송일
                delivery_date = None
                if row[4].value:
                    delivery_date = row[4].value

                # 성분
                ingredient = ''
                if product_current_check is not None:
                    ingredient = product_current_check.ingredient

                # 규격
                standard = ''
                if row[8].value:
                    standard = str(row[8].value)
                    # 불필요 텍스트 제거
                    standard = standard.replace('1kg/', '')
                    standard = standard.replace('국내산/', '')
                    standard = standard.replace('일반/', '')
                    standard = standard.replace('샛별김치/', '')
                    standard = standard.replace('/냉장', '')

                # 품목보고번호
                product_number = ''

                # 포장 1, 3, 5, 10 으로 가능
                # 발주수량을 큰 포장부터 시작하여 최적으로 분류
                # 수정은 차후 excel에서 진행
                try:
                    # 총량
                    totalUnit = row[15].value
                    package = {10: 0, 5: 0, 3: 0, 1: 0}
                    while totalUnit != 0:
                        # 10키로 팩키지 구성하는 경우
                        if totalUnit - 10 == 0 or totalUnit - 10 > 0:
                            package[10] += 1
                            totalUnit -= 10
                        elif totalUnit - 5 == 0 or totalUnit - 5 > 0:
                            package[5] += 1
                            totalUnit -= 5
                        elif totalUnit - 3 == 0 or totalUnit - 3 > 0:
                            package[3] += 1
                            totalUnit -= 3
                        else:
                            package[1] += 1
                            totalUnit -= 1

                    date_manufacture = ''
                    if product_current_check:
                        date_manufacture = product_current_check.date_manufacture

                    date_expiration = ''
                    if product_current_check:
                        date_expiration = product_current_check.date_expiration

                    date_expiration_days = 0
                    try:
                        date_expiration_days = (date_expiration - date_manufacture).days
                    except:
                        if product_current_check:
                            date_expiration_days = product_current_check.date_expiration_days
                            if date_expiration_days is None:
                                date_expiration_days = 0
                        pass

                    expiration_type = '제조일로부터'

                    if date_manufacture:
                        d = product_current_check.date_manufacture
                        date_manufacture = str(d.year) + '년 ' + str(d.month) + '월 ' + str(d.day) + '일'
                    else:
                        date_manufacture = ''
                    if date_expiration:
                        d = product_current_check.date_expiration
                        date_expiration = str(d.year) + '년 ' + str(d.month) + '월 ' + str(d.day) + '일'
                    else:
                        date_expiration = ''

                    if product_current_check:
                        expiration_type = product_current_check.expiration_type

                    if expiration_type == '특정일':
                        date_expiration = date_expiration + '까지'
                    else:
                        if date_expiration_days > 60:
                            date_expiration_days = col_round(date_expiration_days / 30)
                            date_expiration = '제조일로부터 ' + str(date_expiration_days) + '개월까지'
                        else:
                            date_expiration = '제조일로부터 ' + str(date_expiration_days) + '일까지'

                    meal_type = ''
                    if row[25].value:
                        meal_type = row[25].value


                    for box in package:
                        if package[box] > 0:
                            append(LabelPrint_History(date=date, delivery_date=delivery_date,

                                                      product_name=product_current_check.name,
                                                      customer_name=customer_current.delivery_to,
                                                      meal_type=meal_type,

                                                      ingredient=ingredient,
                                                      standard=standard,
                                                      unit_package=box,
                                                      unit=package[box],
                                                      date_manufacture=date_manufacture,
                                                      date_expiration=date_expiration,

                                                      product_number=product_number,
                                                      remarks=remarks,
                                                      created_by=user,
                                                      updated_by=user,
                                                      created_at=now,
                                                      updated_at=now,
                                                      enterprise=enterprise_obj))
                            cnt += 1
                except Exception as e:
                    print('발주서에 수량에 대한 데이터가 없는 경우', e)
                    pass

        # create
        LabelPrint_History.objects.bulk_create(list)

    print('end', datetime.datetime.now() - startedTime)

    return cnt, error_contents

def col_round(x):
  frac = x - math.floor(x)
  if frac < 0.5: return math.floor(x)
  return math.ceil(x)


def LabelPrint_All(request):

    context = {}

    date = request.POST['date']
    customer = request.POST['customer']
    product = request.POST.get('product', None)
    if product == 'null':
        product = None
    enterprise_name = request.POST['enterprise_name']

    user = UserMaster.objects.get(username=request.POST['username'])

    print(date, customer, enterprise_name)

    # 양식 가져오기
    path = os.path.join(BASE_DIR, 'data')
    startedTime = datetime.datetime.now()

    # 양식 복사
    wb, sheet = CopyXlsForm(path)
    # 양식 list
    dict_cell = CopyXlsStyle(sheet)

    list_sheet = []
    appendSheet = list_sheet.append

    cnt = 0
    customer_current = LabelPrint_delivery.objects.get(pk=customer)

    mergedCell = {'C2': 'C2', 'C4': 'C4', 'D4': 'D4'}

    # 대상 조회
    qs_target = LabelPrint_History.objects.filter(
        date=date,
        customer_name=customer_current.delivery_to,
        enterprise__name=enterprise_name
    )
    if product:
        product_current = LabelPrint_product.objects.get(pk=product)
        qs_target = qs_target.filter(product_name=product_current.name)

    for row in qs_target:
        # 수량 만큼 생성
        for i in range(int(row.unit)):
            dict_sheet = {}
            updateSheet = dict_sheet.update

            updateSheet({'qs': row})

            MakeCell(cnt, updateSheet, dict_cell, mergedCell, sheet, appendSheet, dict_sheet)
            cnt += 1

    cnt = 0
    # 데이터 넣기
    for row in list_sheet:
        MakePaysubExcel(row, sheet)
        cnt += 1

    # 페이징 처리
    sheet.print_area = []
    appendPage = sheet.print_area.append

    pageCnt = cnt

    from openpyxl.worksheet.cell_range import CellRange
    mergeLoopCount = int(cnt / 1)
    cnt = 0
    dict_merged = {}
    for mcr in sheet.merged_cells:
        dict_merged.update({cnt: mcr})
        cnt += 1

    for merged in dict_merged:
        for i in range(mergeLoopCount):
            cr = CellRange(dict_merged[merged].coord)
            cr.shift(row_shift=12 + 12 * i)
            sheet.merge_cells(cr.coord)

    for i in range(pageCnt):
        if i == 0:
            continue
        page = math.ceil(i / 1)
        firstArea = 'A' + str((12 * page) - 11)
        appendPage(firstArea + ':' + 'E' + str(page * 12 - 1))

    # 마지막 페이지
    firstArea = 'A' + str(math.ceil(pageCnt / 1) * 12 - 11)
    appendPage(firstArea + ':' + 'E' + str(math.ceil(pageCnt / 1) * 12 - 1))
    # 이 결과가 시트에 완벽하게 적용되려면
    # 결과물을 열어서 페이지 레이아웃을 세로로 절단선에 맞추면 자동으로 모든 페이지의 print_area가 적용 됨

    # 최종 저장
    wb.save(path + '/xlsForm/tmp.xlsx')

    # print(datetime.datetime.now() - startedTime)

    # 파일이름 변경
    context['filename'] = customer_current.delivery_to + '_라벨출력_' + date + '.xlsx'
    context['path'] = path + '/xlsForm/' + context['filename']
    os.rename(path + '/xlsForm/tmp.xlsx', context['path'])

    print(context['path'])

    return JsonResponse(context)


def MakePaysubExcel(row, sheet):

    sheet[row['B2']].value = row['qs'].product_name
    sheet[row['D2']].value = row['qs'].meal_type
    sheet[row['B4']].value = row['qs'].ingredient
    sheet[row['B5']].value = str(row['qs'].unit_package) + 'kg'
    sheet[row['B6']].value = row['qs'].date_manufacture
    sheet[row['B7']].value = row['qs'].customer_name
    sheet[row['E7']].value = row['qs'].date_expiration
    sheet[row['B8']].value = row['qs'].product_number


def MakeCell(cnt, updateSheet, dict_cell, mergedCell, sheet, appendSheet, dict_sheet):
    from copy import copy

    # A ~ E 까지 사용 12칸이 1세트
    # A1 ~ E12 까지의 값 저장
    min_row = 12
    col_array = {0: 'A', 1: 'B', 2: 'C', 3: 'D', 4: 'E'}

    for row in range(min_row):
        currentColIndex = (row + 1) % 12
        target_row = min_row * cnt + row + 1

        for col in col_array:
            # origin
            key_origin = col_array[col] + str(row + 1)
            # target
            key_target = col_array[col] + str(target_row)

            updateSheet({col_array[col] + str(currentColIndex): key_target})

            if str(type(dict_cell[key_origin])) != "<class 'openpyxl.cell.cell.Cell'>":
                continue
            if key_origin in mergedCell:
                continue

            new_cell = sheet.cell(row=target_row, column=col + 1,
                                  value=dict_cell[key_origin].value)

            new_cell.font = copy(dict_cell[key_origin].font)
            new_cell.border = copy(dict_cell[key_origin].border)
            new_cell.fill = copy(dict_cell[key_origin].fill)
            new_cell.number_format = copy(dict_cell[key_origin].number_format)
            new_cell.protection = copy(dict_cell[key_origin].protection)
            new_cell.alignment = copy(dict_cell[key_origin].alignment)
            new_cell.parent = copy(dict_cell[key_origin].parent)
        sheet.row_dimensions[target_row].height = sheet.row_dimensions[currentColIndex].height

    appendSheet(dict_sheet)
    cnt += 1

def CopyXlsStyle(sheet):
    # 양식 list
    dict_cell = {}
    update_cell = dict_cell.update

    # A ~ E 까지 사용 45칸이 1세트
    # A1 ~ K12 까지의 값 저장
    min_row = 12
    col_array = {0: 'A', 1: 'B', 2: 'C', 3: 'D', 4: 'E'}
    for row in range(min_row):
        for col in col_array:
            key = col_array[col] + str(row + 1)
            update_cell({key: sheet[key]})

    return dict_cell


def CopyXlsForm(path):
    # 양식 복사
    baseFiles = {'form_labelPrint.xlsx': 'form_labelPrint.xlsx'}
    filenames = os.listdir(path + '/xlsForm/')
    for filename in filenames:
        if filename not in baseFiles:
            full_filename = os.path.join(path + '/xlsForm/', filename)
            os.remove(full_filename)

    shutil.copy(path + '/xlsForm/form_labelPrint.xlsx', path + '/xlsForm/tmp.xlsx')

    # 파일 열기
    wb = openpyxl.load_workbook(path + '/xlsForm/tmp.xlsx')
    return wb, wb.active
