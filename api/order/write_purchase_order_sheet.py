
from openpyxl import load_workbook


# Write information at the top
def write_info(wb, ws, input_val, file):
    '''
    ws: worksheet
    input_val: dictionary type with (var: value)
    '''
    
    cell_dct = dict(company='G4', charge='G5', tel='G6', 
                    date='O4', po='O5', cpo='O6')
    
    for k,v in input_val.items():
        try:
            cell = cell_dct[k]
            ws[cell] = v
        except KeyError:
            continue
    
    # save the file 
    wb.save(file)
    
    
# Write purchase order information
def write_order(wb, ws, input_val, file):
    '''
    ws: worksheet
    input_val: dictionary type with (var: value)
    '''
    
    cell_dct = dict(no='A10', spec='B9', 
                    qty_num='M10', qty_unit='N10',
                    unit_price='O10', price='P10', 
                    note='B12')
    
    for k,v in input_val.items():
        try:
            cell = cell_dct[k]
            if type(v) == list:
                col = cell[0]
                row = int(cell[1:])
                for i in range(0, len(v)):
                    cell = col + str(row)
                    ws[cell] = v[i]
                    row +=1
            else:
                ws[cell] = v

        except KeyError:
            continue

    wb.save(file)
    
    
# Write note information
def write_note(wb, ws, input_val, file):
    '''
    ws: worksheet
    input_val: dictionary type with (var: value)
    '''
    
    note_dct = dict(payment_terms = 'I32',
                      delivery_terms = 'I33',
                      packing = 'I34',
                      quality_warrenty = 'I35',
                      delivery_location = 'I36',
                     ) 
    
    for k,v in input_val.items():
        try:
            cell = note_dct[k]
            ws[cell] = v
        except KeyError:
            continue

    wb.save(file)
    
    
# Write footer information
def write_footer(wb, ws, input_val, file):
    '''
    ws: worksheet
    input_val: dictionary type with (var: value)
    '''
    
    footer_dct = dict(site = 'A40', manager = 'J40', 
                  email = 'O40', tel2 = 'O41'
                 )
    
    for k,v in input_val.items():
        try:
            cell = footer_dct[k]
            ws[cell] = v
        except KeyError:
            continue

    wb.save(file)


def buildexcel(data, saved, filename='data/order_purchase_template.xlsx'):
    wb = load_workbook(filename)
    ws = wb['Sheet1']

    write_info(wb, ws, data, saved)
    write_order(wb, ws, data, saved)
    write_note(wb, ws, data, saved)
    write_footer(wb, ws, data, saved)
